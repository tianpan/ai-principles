"""
表格结构分析模块

专门负责分析Excel表格结构，识别表头位置、数据区域、多级表头等
"""
import pandas as pd
import numpy as np
import re
from openpyxl import load_workbook
from financial_analysis.logger import log_info, log_debug, log_warning, log_error

class TableAnalyzer:
    """表格结构分析器，用于识别和分析Excel表格的结构"""
    
    def __init__(self):
        """初始化表格结构分析器"""
        # 用于识别日期格式的模式
        self.date_patterns = [
            r'\d{4}[-/年]\d{1,2}[-/月]',  # 年月日期: 2023-01, 2023/01, 2023年01月
            r'\d{4}[-/]Q[1-4]',  # 季度日期: 2023-Q1, 2023/Q2
            r'\d{4}年?第?[一二三四]季度',  # 中文季度: 2023年第一季度, 2023第二季度
            r'\d{4}[-/年]',  # 年份: 2023-, 2023/, 2023年
            r'\d{4}'  # 纯数字年份: 2023
        ]
        
        # 各种可能的表格布局
        self.possible_layouts = [
            {"header_pos": "top", "index_pos": "left"},
            {"header_pos": "left", "index_pos": "top"},
            {"header_pos": "multi-level", "index_pos": "left"},
            {"header_pos": "left", "index_pos": "multi-level"}
        ]
        
        # 日期列识别模式：中文字符或英文字符后紧跟括号中的日期
        self.date_column_patterns = [
            r'.*\((\d{4}[-/]\d{1,2})\)',  # 示例: 资产负债表(2023-12)
            r'.*\((\d{4})\)',  # 示例: 资产(2023)
            r'.*（(\d{4}[-/年]\d{1,2}[-/月]?)）',  # 示例: 资产（2023年12月）
            r'.*（(\d{4})）'  # 示例: 资产（2023）
        ]
        
        # 常见财务表格关键列名识别模式
        self.key_column_patterns = {
            'item_names': [
                r'项目', r'科目', r'名称', r'指标', r'会计科目', 
                r'资产负债表项目', r'损益表项目'
            ],
            'notes': [
                r'注?释', r'附注', r'备注'
            ],
            'unit': [
                r'单位', r'币种', r'金额单位'
            ]
        }
    
    def analyze_table_structure(self, file_path, sheet_name=0):
        """
        分析Excel表格结构
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称或索引
            
        Returns:
            dict: 表格结构分析结果
        """
        log_info(f"开始分析表格结构: {file_path}, 工作表: {sheet_name}", "TableAnalyzer")
        
        # 加载Excel文件
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            log_debug(f"成功加载Excel, 形状: {df.shape}", "TableAnalyzer")
        except Exception as e:
            log_error(f"加载Excel文件失败: {e}", None, "TableAnalyzer")
            return {
                "success": False,
                "error": f"加载Excel文件失败: {str(e)}"
            }
            
        # 分析表格和合并单元格
        merged_cells = self._identify_merged_cells(file_path, sheet_name)
        log_debug(f"检测到{len(merged_cells)}个合并单元格", "TableAnalyzer")
        
        # 检测表头位置
        header_info = self._detect_header_position(df)
        log_debug(f"检测到表头位置: {header_info['header_position']}", "TableAnalyzer")
        
        # 检测日期信息
        date_info = self._extract_dates(df)
        log_debug(f"检测到{len(date_info['dates'])}个日期", "TableAnalyzer")
        
        # 检测单位信息
        unit_info = self._detect_unit(df)
        log_debug(f"检测到单位: {unit_info['unit']}", "TableAnalyzer")
        
        # 识别数据区域
        data_region = self._identify_data_region(df, header_info, date_info)
        log_debug(f"识别数据区域: 行{data_region['start_row']}至{data_region['end_row']}", "TableAnalyzer")
        
        # 返回分析结果
        analysis_result = {
            "success": True,
            "file_path": file_path,
            "sheet_name": sheet_name,
            "table_shape": df.shape,
            "header_info": header_info,
            "date_info": date_info,
            "unit_info": unit_info,
            "data_region": data_region,
            "merged_cells": merged_cells,
            "suggested_parsing_method": self._suggest_parsing_method(header_info, date_info, data_region)
        }
        
        log_info(f"表格结构分析完成", "TableAnalyzer")
        return analysis_result
        
    def _identify_merged_cells(self, file_path, sheet_name=0):
        """
        识别Excel文件中的合并单元格
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称或索引
            
        Returns:
            list: 合并单元格信息列表
        """
        try:
            # 加载工作簿
            workbook = load_workbook(file_path, data_only=True)
            
            # 选择工作表
            if isinstance(sheet_name, int):
                sheet = workbook.worksheets[sheet_name]
            else:
                sheet = workbook[sheet_name]
                
            # 获取合并单元格信息
            merged_cells = []
            for merged_range in sheet.merged_cells.ranges:
                merged_cells.append({
                    "min_row": merged_range.min_row,
                    "max_row": merged_range.max_row,
                    "min_col": merged_range.min_col,
                    "max_col": merged_range.max_col,
                    "value": sheet.cell(merged_range.min_row, merged_range.min_col).value
                })
                
            return merged_cells
        except Exception as e:
            log_warning(f"识别合并单元格时出错: {e}", "TableAnalyzer")
            return []
            
    def _detect_header_position(self, df):
        """
        检测表格的表头位置
        
        Args:
            df: DataFrame数据
            
        Returns:
            dict: 表头位置信息
        """
        # 初始化结果
        header_info = {
            "header_position": "unknown",
            "index_position": "unknown",
            "header_rows": [],
            "header_cols": [],
            "is_multi_level": False,
            "confidence": 0
        }
        
        # 检查是否为常见的横向表头（表头在顶部，项目在左侧）
        horz_header_score = 0
        
        # 检查前5行是否包含日期信息
        date_rows = []
        for i in range(min(5, len(df))):
            row = df.iloc[i]
            date_count = 0
            for val in row:
                if isinstance(val, str) and any(re.search(pattern, val) for pattern in self.date_patterns):
                    date_count += 1
                    
            if date_count >= 2:  # 如果一行中有多个日期，很可能是表头行
                date_rows.append(i)
                horz_header_score += 10 * date_count
                
        # 检查是否为常见的纵向表头（表头在左侧，项目在顶部）
        vert_header_score = 0
        
        # 检查前5列是否包含项目名称关键字
        item_name_cols = []
        for i in range(min(5, len(df.columns))):
            col = df.iloc[:, i]
            item_count = 0
            for val in col:
                if isinstance(val, str) and any(keyword in val for keyword in self.key_column_patterns['item_names']):
                    item_count += 1
                    
            if item_count >= 1:  # 如果一列中有项目名称关键字，可能是纵向表头
                item_name_cols.append(i)
                vert_header_score += 5 * item_count
                
        # 判断表头位置
        if horz_header_score > vert_header_score:
            header_info["header_position"] = "top"
            header_info["index_position"] = "left"
            header_info["header_rows"] = date_rows
            header_info["confidence"] = min(100, horz_header_score)
            
            # 检查是否为多级表头
            if len(date_rows) > 1 and abs(date_rows[0] - date_rows[-1]) <= len(date_rows):
                header_info["is_multi_level"] = True
                
        elif vert_header_score > 0:
            header_info["header_position"] = "left"
            header_info["index_position"] = "top"
            header_info["header_cols"] = item_name_cols
            header_info["confidence"] = min(100, vert_header_score)
            
            # 检查是否为多级表头
            if len(item_name_cols) > 1 and abs(item_name_cols[0] - item_name_cols[-1]) <= len(item_name_cols):
                header_info["is_multi_level"] = True
                
        else:
            # 如果没有明确的表头指示，默认为横向表头
            header_info["header_position"] = "top"
            header_info["index_position"] = "left"
            header_info["header_rows"] = [0]
            header_info["confidence"] = 10
            
        return header_info
        
    def _extract_dates(self, df):
        """
        从DataFrame中提取日期信息
        
        Args:
            df: DataFrame数据
            
        Returns:
            dict: 日期信息
        """
        date_info = {
            "dates": [],
            "date_positions": [],
            "date_format": "unknown"
        }
        
        # 扫描整个DataFrame查找日期
        for i in range(len(df)):
            for j in range(len(df.columns)):
                value = df.iloc[i, j]
                if pd.isna(value):
                    continue
                    
                # 将值转换为字符串
                value_str = str(value)
                
                # 检查是否匹配日期模式
                for pattern in self.date_patterns:
                    if re.search(pattern, value_str):
                        # 提取日期
                        match = re.search(pattern, value_str)
                        date_str = match.group(0)
                        
                        # 标准化日期格式
                        normalized_date = self._normalize_date(date_str)
                        
                        if normalized_date and normalized_date not in date_info["dates"]:
                            date_info["dates"].append(normalized_date)
                            date_info["date_positions"].append((i, j))
                            
                            # 更新日期格式
                            if '月' in date_str or '-' in date_str and len(date_str.split('-')) > 1:
                                date_info["date_format"] = "monthly"
                            elif 'Q' in date_str or '季' in date_str:
                                date_info["date_format"] = "quarterly"
                            else:
                                date_info["date_format"] = "yearly"
                        
                        break
        
        # 排序日期
        date_info["dates"].sort()
        
        return date_info
        
    def _normalize_date(self, date_str):
        """
        标准化日期字符串
        
        Args:
            date_str: 日期字符串
            
        Returns:
            str: 标准化后的日期字符串
        """
        # 处理年月格式
        if re.search(r'\d{4}[-/年]\d{1,2}[-/月]', date_str):
            # 提取年和月
            year = re.search(r'\d{4}', date_str).group(0)
            month = re.search(r'[-/年](\d{1,2})[-/月]?', date_str).group(1)
            return f"{year}-{month.zfill(2)}"
            
        # 处理季度格式
        if re.search(r'\d{4}[-/]Q[1-4]', date_str):
            # 提取年和季度
            year = re.search(r'\d{4}', date_str).group(0)
            quarter = re.search(r'Q([1-4])', date_str).group(1)
            return f"{year}-Q{quarter}"
            
        # 处理中文季度格式
        if re.search(r'\d{4}年?第?[一二三四]季度', date_str):
            # 提取年和季度
            year = re.search(r'\d{4}', date_str).group(0)
            quarter_map = {'一': '1', '二': '2', '三': '3', '四': '4'}
            for cn, num in quarter_map.items():
                if cn in date_str:
                    return f"{year}-Q{num}"
            
        # 处理纯年份格式
        if re.search(r'\d{4}[-/年]', date_str):
            # 提取年
            year = re.search(r'\d{4}', date_str).group(0)
            return year
            
        # 处理纯数字年份
        if re.search(r'^\d{4}$', date_str):
            return date_str
            
        # 无法处理的格式
        return None
        
    def _detect_unit(self, df):
        """
        检测表格中的金额单位
        
        Args:
            df: DataFrame数据
            
        Returns:
            dict: 单位信息
        """
        unit_info = {
            "unit": "元",  # 默认单位为元
            "unit_position": None,
            "confidence": 0
        }
        
        # 单位关键词
        unit_keywords = {
            "元": ["元", "人民币元", "rmb", "cny"],
            "万元": ["万元", "万", "万人民币", "万rmb", "万cny", "w"],
            "亿元": ["亿元", "亿", "亿人民币", "亿rmb", "亿cny", "y"]
        }
        
        # 在前10行和前5列中搜索单位信息
        for i in range(min(10, len(df))):
            for j in range(min(5, len(df.columns))):
                value = df.iloc[i, j]
                if pd.isna(value):
                    continue
                    
                # 将值转换为字符串并转为小写
                value_str = str(value).lower()
                
                # 检查是否包含单位关键词
                for unit, keywords in unit_keywords.items():
                    if any(keyword in value_str for keyword in keywords):
                        unit_info["unit"] = unit
                        unit_info["unit_position"] = (i, j)
                        unit_info["confidence"] = 80
                        return unit_info
                    
                # 检查是否包含"单位"关键词
                if any(keyword in value_str for keyword in self.key_column_patterns['unit']):
                    # 如果同一单元格中包含单位关键词
                    for unit, keywords in unit_keywords.items():
                        if any(keyword in value_str for keyword in keywords):
                            unit_info["unit"] = unit
                            unit_info["unit_position"] = (i, j)
                            unit_info["confidence"] = 90
                            return unit_info
                    
                    # 检查相邻单元格
                    neighbors = [
                        (i, j+1), (i+1, j), (i, j-1), (i-1, j)
                    ]
                    
                    for ni, nj in neighbors:
                        if 0 <= ni < len(df) and 0 <= nj < len(df.columns):
                            neighbor_value = df.iloc[ni, nj]
                            if pd.isna(neighbor_value):
                                continue
                                
                            neighbor_str = str(neighbor_value).lower()
                            for unit, keywords in unit_keywords.items():
                                if any(keyword in neighbor_str for keyword in keywords):
                                    unit_info["unit"] = unit
                                    unit_info["unit_position"] = (ni, nj)
                                    unit_info["confidence"] = 85
                                    return unit_info
        
        # 根据数值大小推断单位
        num_values = []
        for i in range(len(df)):
            for j in range(len(df.columns)):
                value = df.iloc[i, j]
                if isinstance(value, (int, float)) and not pd.isna(value):
                    num_values.append(value)
                    
        if num_values:
            avg_value = np.mean([abs(v) for v in num_values if v != 0])
            if avg_value > 1e10:  # 平均值很大，可能是"元"
                unit_info["unit"] = "元"
                unit_info["confidence"] = 30
            elif avg_value > 1e6:  # 平均值适中，可能是"万元"
                unit_info["unit"] = "万元"
                unit_info["confidence"] = 40
            else:  # 平均值较小，可能是"亿元"
                unit_info["unit"] = "亿元"
                unit_info["confidence"] = 20
                
        return unit_info
        
    def _identify_data_region(self, df, header_info, date_info):
        """
        识别表格中的数据区域
        
        Args:
            df: DataFrame数据
            header_info: 表头信息
            date_info: 日期信息
            
        Returns:
            dict: 数据区域信息
        """
        data_region = {
            "start_row": 0,
            "end_row": len(df) - 1,
            "start_col": 0,
            "end_col": len(df.columns) - 1,
            "confidence": 50
        }
        
        # 根据表头位置确定数据区域
        if header_info["header_position"] == "top":
            # 表头在顶部，数据区域从表头下方开始
            if header_info["header_rows"]:
                data_region["start_row"] = max(header_info["header_rows"]) + 1
                data_region["confidence"] = 70
                
        elif header_info["header_position"] == "left":
            # 表头在左侧，数据区域从表头右侧开始
            if header_info["header_cols"]:
                data_region["start_col"] = max(header_info["header_cols"]) + 1
                data_region["confidence"] = 70
                
        # 使用日期位置进一步调整数据区域
        if date_info["date_positions"]:
            date_rows = [pos[0] for pos in date_info["date_positions"]]
            date_cols = [pos[1] for pos in date_info["date_positions"]]
            
            if header_info["header_position"] == "top":
                data_region["start_row"] = max(data_region["start_row"], max(date_rows) + 1)
                data_region["confidence"] = 80
                
            elif header_info["header_position"] == "left":
                data_region["start_col"] = max(data_region["start_col"], max(date_cols) + 1)
                data_region["confidence"] = 80
                
        # 识别数据结束位置
        # 从后向前扫描，找到最后一个非空行
        for i in range(len(df) - 1, data_region["start_row"] - 1, -1):
            row = df.iloc[i]
            if not row.isna().all() and not all(str(x).strip() == '' for x in row if not pd.isna(x)):
                data_region["end_row"] = i
                break
                
        # 从后向前扫描，找到最后一个非空列
        for j in range(len(df.columns) - 1, data_region["start_col"] - 1, -1):
            col = df.iloc[:, j]
            if not col.isna().all() and not all(str(x).strip() == '' for x in col if not pd.isna(x)):
                data_region["end_col"] = j
                break
                
        return data_region
        
    def _suggest_parsing_method(self, header_info, date_info, data_region):
        """
        根据分析结果建议最佳的解析方法
        
        Args:
            header_info: 表头信息
            date_info: 日期信息
            data_region: 数据区域信息
            
        Returns:
            dict: 解析方法建议
        """
        method = {
            "name": "standard",
            "params": {},
            "confidence": 50
        }
        
        # 基于表头位置和日期格式确定解析方法
        if header_info["header_position"] == "top" and header_info["is_multi_level"]:
            method["name"] = "multi_level_header_top"
            method["params"]["header_rows"] = header_info["header_rows"]
            method["confidence"] = 70
            
        elif header_info["header_position"] == "left" and header_info["is_multi_level"]:
            method["name"] = "multi_level_header_left"
            method["params"]["header_cols"] = header_info["header_cols"]
            method["confidence"] = 70
            
        elif header_info["header_position"] == "top":
            method["name"] = "standard_header_top"
            method["params"]["header_row"] = max(header_info["header_rows"]) if header_info["header_rows"] else 0
            method["confidence"] = 80
            
        elif header_info["header_position"] == "left":
            method["name"] = "standard_header_left"
            method["params"]["header_col"] = max(header_info["header_cols"]) if header_info["header_cols"] else 0
            method["confidence"] = 80
            
        # 添加日期信息
        method["params"]["date_format"] = date_info["date_format"]
        method["params"]["dates"] = date_info["dates"]
        
        # 添加数据区域信息
        method["params"]["data_start_row"] = data_region["start_row"]
        method["params"]["data_end_row"] = data_region["end_row"]
        method["params"]["data_start_col"] = data_region["start_col"]
        method["params"]["data_end_col"] = data_region["end_col"]
        
        return method
        
    def preprocess_table(self, df, analysis_result):
        """
        根据分析结果预处理表格
        
        Args:
            df: 原始DataFrame
            analysis_result: 表格分析结果
            
        Returns:
            DataFrame: 预处理后的DataFrame
        """
        # 获取解析方法
        method = analysis_result["suggested_parsing_method"]
        
        # 根据不同解析方法处理表格
        if method["name"] == "multi_level_header_top":
            # 处理顶部多级表头
            header_rows = method["params"]["header_rows"]
            df_processed = df.copy()
            
            # 使用多级表头
            if len(header_rows) > 1:
                # 获取表头行
                headers = [df.iloc[i] for i in header_rows]
                
                # 构建多级索引
                multi_header = []
                for i in range(len(df.columns)):
                    header_values = [str(headers[j][i]).strip() if not pd.isna(headers[j][i]) else '' for j in range(len(headers))]
                    multi_header.append(tuple(header_values))
                    
                # 设置多级列索引
                df_processed.columns = pd.MultiIndex.from_tuples(multi_header)
                
                # 去除表头行
                df_processed = df_processed.iloc[max(header_rows) + 1:].reset_index(drop=True)
                
            return df_processed
            
        elif method["name"] == "standard_header_top":
            # 处理标准顶部表头
            header_row = method["params"]["header_row"]
            df_processed = df.copy()
            
            # 使用表头行作为列名
            df_processed.columns = df.iloc[header_row].values
            
            # 去除表头行
            df_processed = df_processed.iloc[header_row + 1:].reset_index(drop=True)
            
            return df_processed
            
        elif method["name"] == "standard_header_left":
            # 处理左侧表头
            header_col = method["params"]["header_col"]
            df_processed = df.copy()
            
            # 转置DataFrame
            df_processed = df_processed.T
            
            # 使用表头列作为列名
            df_processed.columns = df.iloc[:, header_col].values
            
            # 去除表头列
            df_processed = df_processed.iloc[header_col + 1:].reset_index(drop=True)
            
            return df_processed
            
        elif method["name"] == "multi_level_header_left":
            # 处理左侧多级表头
            header_cols = method["params"]["header_cols"]
            df_processed = df.copy()
            
            # 转置DataFrame
            df_processed = df_processed.T
            
            # 使用多级表头
            if len(header_cols) > 1:
                # 获取表头列
                headers = [df.iloc[:, i] for i in header_cols]
                
                # 构建多级索引
                multi_header = []
                for i in range(len(df)):
                    header_values = [str(headers[j][i]).strip() if not pd.isna(headers[j][i]) else '' for j in range(len(headers))]
                    multi_header.append(tuple(header_values))
                    
                # 设置多级列索引
                df_processed.columns = pd.MultiIndex.from_tuples(multi_header)
                
                # 去除表头列
                df_processed = df_processed.iloc[max(header_cols) + 1:].reset_index(drop=True)
                
            return df_processed
            
        else:
            # 默认处理方法
            return df
            
    def extract_key_items(self, df, item_patterns):
        """
        从DataFrame中提取关键项目
        
        Args:
            df: DataFrame
            item_patterns: 项目匹配模式
            
        Returns:
            dict: 提取的关键项目
        """
        extracted_items = {}
        
        # 检查DataFrame的每一行
        for idx, row in df.iterrows():
            # 获取第一列作为项目名称
            item_name = row.iloc[0]
            if pd.isna(item_name) or not isinstance(item_name, str):
                continue
                
            # 检查是否匹配任何项目模式
            for category, patterns in item_patterns.items():
                if any(pattern in item_name for pattern in patterns):
                    # 获取该行的所有数值
                    values = {}
                    for i, col in enumerate(df.columns[1:], 1):
                        if isinstance(col, tuple):
                            col_name = '_'.join(str(x) for x in col if x)
                        else:
                            col_name = str(col)
                            
                        value = row.iloc[i]
                        if not pd.isna(value) and (isinstance(value, (int, float)) or (isinstance(value, str) and self._is_numeric(value))):
                            values[col_name] = self._parse_numeric(value)
                            
                    # 将提取的项目添加到结果中
                    key = f"{category}_{item_name}"
                    extracted_items[key] = values
                    break
                    
        return extracted_items
        
    def _is_numeric(self, value):
        """
        检查字符串是否可以转换为数值
        
        Args:
            value: 要检查的值
            
        Returns:
            bool: 是否可以转换为数值
        """
        if isinstance(value, (int, float)):
            return True
            
        if not isinstance(value, str):
            return False
            
        # 清理字符串
        value = value.replace(',', '').replace(' ', '')
        
        # 处理括号表示的负数
        if '(' in value and ')' in value:
            value = value.replace('(', '-').replace(')', '')
            
        # 尝试转换为浮点数
        try:
            float(value)
            return True
        except ValueError:
            return False
            
    def _parse_numeric(self, value):
        """
        解析并转换数值
        
        Args:
            value: 要解析的值
            
        Returns:
            float: 解析后的数值
        """
        if isinstance(value, (int, float)):
            return float(value)
            
        if not isinstance(value, str):
            return np.nan
            
        # 清理字符串
        value = value.replace(',', '').replace(' ', '')
        
        # 处理括号表示的负数
        if '(' in value and ')' in value:
            value = value.replace('(', '-').replace(')', '')
            
        # 尝试转换为浮点数
        try:
            return float(value)
        except ValueError:
            return np.nan 