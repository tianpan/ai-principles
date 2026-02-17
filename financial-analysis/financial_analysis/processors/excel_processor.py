"""
Excel数据处理模块

用于处理资产负债表和损益表的Excel文件，提取关键财务数据。
"""
import pandas as pd
import numpy as np
import re
import os
import warnings
from openpyxl import load_workbook
from .template_store import TemplateStore
from .data_validator import DataValidator
from .excel_capture_logger import ExcelCaptureLogger
# 修改为绝对导入
from financial_analysis.logger import log_info, log_debug, log_warning, log_error


class ExcelProcessor:
    """用于处理财务Excel文件的处理器类"""

    def __init__(self, balance_sheet_path, income_statement_path, template_dir=None):
        """
        初始化Excel处理器
        
        Args:
            balance_sheet_path: 资产负债表文件路径
            income_statement_path: 损益表文件路径
            template_dir: 模板存储目录
        """
        log_info(f"初始化Excel处理器，资产负债表: {balance_sheet_path}, 损益表: {income_statement_path}", "ExcelProcessor")
        self.balance_sheet_path = balance_sheet_path
        self.income_statement_path = income_statement_path
        self.company_name = self._extract_company_name()
        log_debug(f"提取的公司名称: {self.company_name}", "ExcelProcessor")
        
        # 初始化模板存储和数据验证器
        self.template_store = TemplateStore(template_dir)
        self.data_validator = DataValidator()
        
        # 初始化Excel数据捕获日志器
        self.excel_logger = ExcelCaptureLogger()
        
        # 记录处理日志
        self.processing_logs = []
        
        # 表格结构识别选项
        self.table_layouts = [
            {"header_position": "top", "date_position": "top"},
            {"header_position": "left", "date_position": "top"},
            {"header_position": "top", "date_position": "left"},
            {"header_position": "left", "date_position": "left"}
        ]
        
        # 扩展关键字匹配模式
        self.balance_sheet_patterns = {
            'start': [
                '资产', '资产负债表', 'BS1', '流动资产', '货币资金', '应收账款',
                '资产类', '流动资产类', '资产项目', '资产部分'
            ],
            'end': [
                '负债和所有者权益', '负债及股东权益合计', '负债和股东权益',
                '负债股东权益总计', '负债和所有者权益总计'
            ],
            'assets': [
                '流动资产', '非流动资产', '资产总计', '资产合计', '总资产',
                '货币资金', '应收账款', '存货', '固定资产', '无形资产'
            ],
            'liabilities': [
                '流动负债', '非流动负债', '负债合计', '总负债',
                '短期借款', '应付账款', '长期借款', '应付债券'
            ],
            'equity': [
                '所有者权益', '股东权益', '实收资本', '资本公积',
                '盈余公积', '未分配利润', '少数股东权益'
            ]
        }
        
        self.income_statement_patterns = {
            'start': [
                '营业收入', '损益表', '利润表', '主营业务收入',
                '收入', '收入项目', '损益项目'
            ],
            'end': [
                '净利润', '本年利润', '归属于母公司', '利润总额',
                '本期利润', '税后净利', '归属于母公司所有者的净利润'
            ],
            'revenue': [
                '营业收入', '主营业务收入', '其他业务收入',
                '总收入', '收入', '销售收入'
            ],
            'costs': [
                '营业成本', '主营业务成本', '其他业务成本',
                '销售费用', '管理费用', '财务费用', '研发费用'
            ],
            'profit': [
                '营业利润', '利润总额', '净利润', '毛利',
                '毛利润', '税前利润', '税后利润'
            ]
        }
    
    def _log_processing_info(self, message):
        """
        记录处理信息
        
        Args:
            message: 处理信息
        """
        self.processing_logs.append({"type": "info", "message": message})
        log_info(message, "ExcelProcessor")
        
    def _log_processing_warning(self, message):
        """
        记录处理警告
        
        Args:
            message: 处理警告
        """
        self.processing_logs.append({"type": "warning", "message": message})
        log_warning(message, "ExcelProcessor")
        
    def _log_processing_error(self, message):
        """
        记录处理错误
        
        Args:
            message: 处理错误
        """
        self.processing_logs.append({"type": "error", "message": message})
        log_error(message, None, "ExcelProcessor")
    
    def _get_processing_log_report(self):
        """
        获取处理日志报告
        
        Returns:
            dict: 处理日志报告
        """
        log_count = {
            "info": 0,
            "warning": 0,
            "error": 0
        }
        
        for log in self.processing_logs:
            log_count[log["type"]] += 1
            
        return {
            "logs": self.processing_logs,
            "counts": log_count,
            "total": len(self.processing_logs)
        }
    
    def _extract_company_name(self):
        """
        从文件名提取公司名称
        
        Returns:
            str: 公司名称
        """
        # 从文件名提取公司名，假设文件名包含公司名
        file_name = os.path.basename(self.balance_sheet_path)
        # 去除扩展名和常见前缀
        name = os.path.splitext(file_name)[0]
        name = re.sub(r'资产负债表|balance_sheet|bs', '', name, flags=re.IGNORECASE)
        
        if not name:
            # 如果无法从资产负债表获取，尝试从损益表获取
            file_name = os.path.basename(self.income_statement_path)
            name = os.path.splitext(file_name)[0]
            name = re.sub(r'损益表|income_statement|is', '', name, flags=re.IGNORECASE)
        
        # 如果仍然无法获取，使用默认名称
        return name.strip() if name.strip() else "未命名企业"
    
    def _find_data_range(self, df, keywords):
        """
        在DataFrame中找到包含关键字的行索引范围
        
        Args:
            df: 数据框
            keywords: 关键字列表
            
        Returns:
            tuple: (开始行索引, 结束行索引)
        """
        start_idx = None
        end_idx = None
        
        for idx, row in df.iterrows():
            row_str = ' '.join([str(x) for x in row if pd.notna(x)])
            
            # 查找起始行
            if start_idx is None:
                for keyword in keywords['start']:
                    if keyword in row_str:
                        start_idx = idx
                        break
                        
            # 查找结束行
            elif end_idx is None:
                for keyword in keywords['end']:
                    if keyword in row_str:
                        end_idx = idx
                        break
        
        # 如果没找到结束行，使用最后一行
        if start_idx is not None and end_idx is None:
            end_idx = df.index[-1]
            
        return start_idx, end_idx
    
    def _clean_numeric_data(self, value):
        """
        清理和转换数值数据
        
        Args:
            value: 原始值
            
        Returns:
            float: 清理后的数值
        """
        if pd.isna(value):
            return np.nan
            
        # 如果已经是数值类型，直接返回
        if isinstance(value, (int, float)):
            return float(value)
            
        # 转换为字符串
        value_str = str(value)
        
        # 清理字符串中的非数值字符
        # 去除千位分隔符
        value_str = value_str.replace(',', '')
        
        # 处理括号表示的负数
        if '(' in value_str and ')' in value_str:
            value_str = '-' + value_str.replace('(', '').replace(')', '')
            
        # 去除前后空白字符
        value_str = value_str.strip()
        
        # 尝试转换为浮点数
        try:
            return float(value_str)
        except ValueError:
            return np.nan
    
    def process_balance_sheet(self):
        """
        处理资产负债表
        
        Returns:
            dict: 处理结果
        """
        self._log_processing_info(f"开始处理资产负债表：{self.balance_sheet_path}")
        
        try:
            # 使用Excel数据捕获日志记录文件信息
            workbook = load_workbook(self.balance_sheet_path, data_only=True)
            self.excel_logger.log_file_info(
                self.balance_sheet_path, 
                sheet_name=workbook.sheetnames[0] if workbook.sheetnames else None,
                sheet_count=len(workbook.sheetnames)
            )
            
            # 尝试不同方式读取Excel
            try:
                df = pd.read_excel(self.balance_sheet_path)
                self._log_processing_info(f"成功读取Excel文件，数据形状: {df.shape}")
                self.excel_logger.log_data_range(df.shape)
            except Exception as e:
                self._log_processing_error(f"读取Excel文件失败: {e}")
                return {"success": False, "error": f"读取Excel文件失败: {str(e)}"}

            # 检测单位
            unit = self._detect_unit_scale(df)
            self._log_processing_info(f"检测到单位: {unit}")
            
            # 记录有效数据范围
            valid_range = self._find_data_range(df, self.balance_sheet_patterns)
            self.excel_logger.log_data_range(
                df.shape,
                valid_range={"start_row": valid_range[0] if valid_range[0] is not None else 0, 
                            "end_row": valid_range[1] if valid_range[1] is not None else df.shape[0]}
            )
            
            # 检测表头位置
            layout = self._detect_header_position(df)
            self._log_processing_info(f"检测到的表格布局: 表头位置={layout['header_position']}")
            
            # 处理多级表头
            parsed_df = self._parse_multi_level_header(df)
            
            # 识别合并单元格
            merged_cells = self._identify_merged_cells(self.balance_sheet_path)
            if merged_cells:
                self._log_processing_info(f"检测到 {len(merged_cells)} 个合并单元格")
                
            # 尝试多种读取方式
            result = None
            successful_patterns = []
            
            # 标准方法
            result, successful_patterns = self._try_process_balance_sheet(parsed_df, self.balance_sheet_patterns, unit)
            
            # 如果标准方法失败，尝试跳过前几行
            if not result or not result['assets']:
                self._log_processing_info("标准方法未能识别到资产负债表数据，尝试跳过前几行")
                for skip_rows in [1, 2, 3, 5, 7]:
                    self._log_processing_info(f"尝试跳过前 {skip_rows} 行")
                    try:
                        alt_df = pd.read_excel(self.balance_sheet_path, header=None, skiprows=skip_rows)
                        alt_parsed_df = self._parse_multi_level_header(alt_df)
                        result, successful_patterns = self._try_process_balance_sheet(alt_parsed_df, self.balance_sheet_patterns, unit)
                        if result and result['assets']:
                            self._log_processing_info(f"成功读取资产负债表，跳过前 {skip_rows} 行")
                            break
                    except Exception as e:
                        self._log_processing_warning(f"尝试跳过前 {skip_rows} 行时出错: {e}")
                        
            # 如果仍然失败，尝试自定义表头
            if not result or not result['assets']:
                self._log_processing_info("尝试使用自定义表头读取")
                try:
                    custom_df = pd.read_excel(self.balance_sheet_path, header=None)
                    result, successful_patterns = self._try_process_balance_sheet_custom(custom_df, unit)
                except Exception as e:
                    self._log_processing_warning(f"尝试使用自定义表头时出错: {e}")
            
            if result and result['assets']:
                # 验证处理结果
                validation_result = self.data_validator.validate_balance_sheet(result)
                if not validation_result['is_valid']:
                    self._log_processing_warning("资产负债表数据验证未通过")
                    for warning in validation_result.get('warnings', []):
                        self._log_processing_warning(warning)
                
                # 保存成功的模板
                if successful_patterns:
                    try:
                        template_id = self.template_store.save_balance_sheet_template(parsed_df, successful_patterns)
                        self._log_processing_info(f"成功保存资产负债表模板，ID: {template_id}")
                    except Exception as e:
                        self._log_processing_warning(f"保存模板时出错: {e}")
                
                return result
            else:
                self._log_processing_error("所有方法均未能成功识别资产负债表数据")
                return {
                    'assets': {},
                    'liabilities': {},
                    'equity': {},
                    'dates': []
                }
                
        except Exception as e:
            self._log_processing_error(f"处理资产负债表时出错: {str(e)}")
            raise
    
    def _process_balance_sheet_with_template(self, df, template_suggestion):
        """
        使用模板处理资产负债表
        
        Args:
            df: 输入的DataFrame
            template_suggestion: 模板建议
            
        Returns:
            dict: 处理后的资产负债表数据
        """
        patterns = template_suggestion.get('patterns', self.balance_sheet_patterns)
        
        # 检测单位
        unit = self._detect_unit_scale(df)
        
        # 使用模板提供的模式尝试识别
        return self._try_process_balance_sheet(df, {'start': patterns, 'end': patterns}, unit)[0]
    
    def _try_process_balance_sheet(self, df, keywords, unit):
        """
        尝试处理资产负债表
        
        Args:
            df: 输入的DataFrame
            keywords: 关键字字典
            unit: 数值单位
            
        Returns:
            tuple: (处理结果, 成功的匹配模式)
        """
        start_idx, end_idx = self._find_data_range(df, keywords)
        
        if start_idx is None:
            return None, []
            
        # 提取有效数据范围
        data_df = df.iloc[start_idx:end_idx+1]
        
        # 识别日期列
        date_columns = []
        for col_idx, col in enumerate(df.columns):
            col_values = df[col].astype(str)
            date_matches = col_values.str.contains(r'\d{4}[-/]\d{1,2}', regex=True)
            if date_matches.any():
                date_columns.append(col_idx)  # 存储列索引而非列名
        
        # 如果没有找到日期列，尝试识别数值列
        if not date_columns:
            numeric_columns = []
            for col_idx, col in enumerate(df.columns):
                if pd.notna(df[col]).any() and df[col].apply(lambda x: isinstance(x, (int, float)) or 
                                 (isinstance(x, str) and re.match(r'^-?\d+\.?\d*$', str(x).replace(',', '')))).any():
                    numeric_columns.append(col_idx)  # 存储列索引而非列名
            date_columns = numeric_columns[:3]  # 假设前三个数值列为日期列
        
        # 提取资产、负债和所有者权益
        assets = {}
        liabilities = {}
        equity = {}
        
        # 提取日期
        dates = []
        for col_idx in date_columns:
            if col_idx < len(df.columns):
                date_value = df.iloc[0, col_idx]
                if pd.notna(date_value):
                    dates.append(str(date_value))
                else:
                    dates.append(f"Column {col_idx}")
        
        # 如果没有从列中找到日期，尝试从表格内容中提取
        if not dates or all(d.startswith("Column") for d in dates):
            extracted_dates = self._extract_dates_from_df(df)
            if extracted_dates:
                dates = extracted_dates
                
        # 确保日期格式一致
        clean_dates = []
        for date in dates:
            match = re.search(r'(\d{4})[-/](\d{1,2})', str(date))
            if match:
                year, month = match.groups()
                clean_dates.append(f"{year}-{month.zfill(2)}")
            else:
                clean_dates.append(date)
        
        dates = clean_dates
        
        # 记录成功匹配的模式
        successful_patterns = []
        
        # 处理数据行
        for idx, row in data_df.iterrows():
            item_name = None
            # 查找第一列作为项目名称
            if len(row) > 0:
                first_col_value = row.iloc[0]  # 使用位置索引
                if pd.notna(first_col_value) and isinstance(first_col_value, str) and len(first_col_value.strip()) > 0:
                    item_name = first_col_value.strip()
                
            if not item_name:
                continue
            
            # 记录成功匹配的项目名称
            if any(keyword in item_name for keyword in keywords.get('assets', [])):
                successful_patterns.append(item_name)
            elif any(keyword in item_name for keyword in keywords.get('liabilities', [])):
                successful_patterns.append(item_name)
            elif any(keyword in item_name for keyword in keywords.get('equity', [])):
                successful_patterns.append(item_name)
                
            # 根据关键字确定是资产、负债还是所有者权益
            row_dict = {}
            for i, col_idx in enumerate(date_columns):
                if i < len(dates) and col_idx < len(row):
                    value = row.iloc[col_idx]  # 使用位置索引
                    # 清理和转换数值，同时统一单位
                    cleaned_value = self._clean_numeric_data(value)
                    if not np.isnan(cleaned_value):
                        normalized_value = self._normalize_unit_scale(cleaned_value, unit)
                        row_dict[dates[i]] = normalized_value
            
            if any(keyword in item_name for keyword in ['资产', 'BS1', '流动资产', '非流动资产']):
                assets[item_name] = row_dict
            elif any(keyword in item_name for keyword in ['负债', 'BS3', '流动负债', '非流动负债']):
                liabilities[item_name] = row_dict
            elif any(keyword in item_name for keyword in ['所有者权益', '股东权益', 'BS5', '实收资本', '资本公积']):
                equity[item_name] = row_dict
        
        return {
            'assets': assets,
            'liabilities': liabilities,
            'equity': equity,
            'dates': dates,
            'unit': unit
        }, successful_patterns
    
    def _try_process_balance_sheet_custom(self, df, unit):
        """
        使用自定义方法处理复杂的资产负债表
        
        Args:
            df: 输入的DataFrame
            unit: 数值单位
            
        Returns:
            tuple: (处理结果, 成功的匹配模式)
        """
        assets = {}
        liabilities = {}
        equity = {}
        successful_patterns = []
        
        # 提取日期信息
        dates = self._extract_dates_from_df(df)
        
        # 如果没有找到日期，尝试使用列索引
        if not dates:
            numeric_cols = []
            for i, col in enumerate(df.columns):
                if df[col].apply(lambda x: isinstance(x, (int, float)) or 
                              (isinstance(x, str) and re.match(r'^-?\d+\.?\d*$', str(x).replace(',', '')))).any():
                    numeric_cols.append(i)
            
            if numeric_cols:
                dates = [f"Column {col}" for col in numeric_cols[:3]]
        
        # 查找资产、负债和所有者权益的行
        for idx, row in df.iterrows():
            item_found = False
            item_name = None
            
            # 查找第一列是否包含项目名称
            if len(df.columns) > 0:
                first_col_value = row.iloc[0]  # 使用位置索引
                if pd.notna(first_col_value) and isinstance(first_col_value, str):
                    item_name = first_col_value.strip()
                    
                    # 检查是否为资产、负债或所有者权益项目
                    if any(keyword in item_name for keyword in 
                          self.balance_sheet_patterns['assets'] + 
                          self.balance_sheet_patterns['liabilities'] + 
                          self.balance_sheet_patterns['equity']):
                        item_found = True
                        successful_patterns.append(item_name)
            
            if not item_found or not item_name:
                continue
                
            # 提取数值
            row_dict = {}
            date_cols = []
            
            # 如果有日期，使用日期作为键
            if dates:
                date_cols = list(range(1, min(len(df.columns), len(dates) + 1)))
                
                for i, col_idx in enumerate(date_cols):
                    if i < len(dates) and col_idx < len(df.columns):
                        value = row.iloc[col_idx]  # 使用位置索引
                        # 清理和转换数值
                        cleaned_value = self._clean_numeric_data(value)
                        if not np.isnan(cleaned_value):
                            normalized_value = self._normalize_unit_scale(cleaned_value, unit)
                            row_dict[dates[i]] = normalized_value
            
            # 根据项目名称分类
            if any(keyword in item_name for keyword in self.balance_sheet_patterns['assets']):
                assets[item_name] = row_dict
            elif any(keyword in item_name for keyword in self.balance_sheet_patterns['liabilities']):
                liabilities[item_name] = row_dict
            elif any(keyword in item_name for keyword in self.balance_sheet_patterns['equity']):
                equity[item_name] = row_dict
        
        return {
            'assets': assets,
            'liabilities': liabilities,
            'equity': equity,
            'dates': dates,
            'unit': unit
        }, successful_patterns
    
    def _detect_unit_scale(self, df):
        """
        检测数值的单位规模（元、万元、亿元等）
        
        Args:
            df: 输入的DataFrame
            
        Returns:
            str: 检测到的单位
        """
        # 查找单位关键词
        for idx, row in df.iterrows():
            row_str = ' '.join([str(x) for x in row if pd.notna(x)])
            
            # 检查单位指示
            if re.search(r'单位[：:]\s*万元', row_str):
                return "万元"
            elif re.search(r'单位[：:]\s*亿元', row_str):
                return "亿元"
            elif re.search(r'单位[：:]\s*元', row_str):
                return "元"
            elif re.search(r'单位[：:]\s*千元', row_str):
                return "千元"
        
        # 如果没有找到明确的单位指示，尝试通过数值大小推断
        numeric_values = []
        for col in df.columns:
            try:
                numeric_values.extend(df[col].dropna().apply(lambda x: float(str(x).replace(',', '')) 
                                                            if isinstance(x, (str, int, float)) and re.match(r'^-?\d+\.?\d*$', str(x).replace(',', '')) 
                                                            else np.nan).dropna().tolist())
            except:
                continue
                
        if numeric_values:
            avg_magnitude = np.mean([abs(x) for x in numeric_values if not np.isnan(x)])
            
            if avg_magnitude < 1000:
                return "元"
            elif avg_magnitude < 10000000:
                return "万元"
            else:
                return "亿元"
                
        return "元"  # 默认单位
        
    def _normalize_unit_scale(self, value, source_unit, target_unit="元"):
        """
        统一数值单位
        
        Args:
            value: 原始数值
            source_unit: 原始单位
            target_unit: 目标单位
            
        Returns:
            float: 转换后的数值
        """
        if pd.isna(value):
            return np.nan
            
        # 定义转换因子
        unit_factors = {
            "元": 1,
            "千元": 1000,
            "万元": 10000,
            "亿元": 100000000
        }
        
        if source_unit not in unit_factors or target_unit not in unit_factors:
            return value
            
        # 转换到目标单位
        return value * (unit_factors[source_unit] / unit_factors[target_unit])
    
    def _identify_merged_cells(self, file_path):
        """
        识别Excel文件中的合并单元格
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            list: 合并单元格信息列表
        """
        try:
            # 加载工作簿
            workbook = load_workbook(file_path, data_only=True)
            
            # 选择第一个工作表
            sheet = workbook.worksheets[0]
            
            # 获取合并单元格信息
            merged_cells = []
            for merged_range in sheet.merged_cells.ranges:
                # 创建单元格范围字符串（例如："A1:B2"）
                min_col_letter = sheet.cell(1, merged_range.min_col).column_letter
                max_col_letter = sheet.cell(1, merged_range.max_col).column_letter
                range_str = f"{min_col_letter}{merged_range.min_row}:{max_col_letter}{merged_range.max_row}"
                
                # 获取合并单元格的值
                value = sheet.cell(merged_range.min_row, merged_range.min_col).value
                
                # 存储合并单元格信息
                merged_cell_info = {
                    "min_row": merged_range.min_row,
                    "max_row": merged_range.max_row,
                    "min_col": merged_range.min_col,
                    "max_col": merged_range.max_col,
                    "value": value
                }
                merged_cells.append(merged_cell_info)
                
                # 记录合并单元格日志
                self.excel_logger.log_merged_cell(
                    range_str,
                    value,
                    processing_result="identified"
                )
            
            self._log_processing_info(f"检测到{len(merged_cells)}个合并单元格")
            return merged_cells
        except Exception as e:
            self._log_processing_warning(f"识别合并单元格时出错: {e}")
            return []
    
    def _detect_header_position(self, df):
        """
        检测表格的表头位置
        
        Args:
            df: DataFrame数据
            
        Returns:
            dict: 表头位置信息
        """
        # 初始化结果
        header_info = {
            "header_position": "unknown",
            "header_rows": [],
            "is_multi_level": False,
            "confidence": 0
        }
        
        # 可能的表头关键词
        header_keywords = [
            "项目", "科目", "资产", "负债", "所有者权益",
            "收入", "成本", "费用", "利润", "item", "科目名称"
        ]
        
        # 可能的日期模式
        date_patterns = [
            r'\d{4}[-/年]\d{1,2}[-/月]',
            r'\d{4}[-/]Q[1-4]',
            r'\d{4}年?第?[一二三四]季度',
            r'\d{4}[-/年]',
            r'\d{4}'
        ]
        
        # 检查表头是否在顶部
        header_in_top = False
        for col in df.columns:
            col_values = df[col].astype(str)
            if col_values.str.contains('|'.join(self.balance_sheet_patterns['assets'] + 
                                              self.balance_sheet_patterns['liabilities'] + 
                                              self.balance_sheet_patterns['equity'])).any():
                header_in_top = False
                break
                
            if col_values.str.contains('|'.join(self.income_statement_patterns['revenue'] + 
                                             self.income_statement_patterns['costs'] + 
                                             self.income_statement_patterns['profit'])).any():
                header_in_top = False
                break
        
        # 检查日期是否在顶部
        date_in_top = False
        for col in df.columns:
            col_values = df[col].astype(str)
            if col_values.str.contains(r'\d{4}[-/]\d{1,2}', regex=True).any():
                date_in_top = True
                break
                
        # 检查日期是否在左侧
        date_in_left = False
        for idx, row in df.iterrows():
            row_str = ' '.join([str(x) for x in row if pd.notna(x)])
            if re.search(r'\d{4}[-/]\d{1,2}', row_str):
                date_in_left = True
                break
        
        # 定义安全的关键词检查函数，处理浮点数情况
        def safe_contains_keyword(x, keywords):
            try:
                if isinstance(x, (int, float)):
                    # 对于数字类型，直接转为字符串并检查
                    return any(keyword in str(x) for keyword in keywords)
                elif isinstance(x, str):
                    # 对于字符串，直接检查
                    return any(keyword in x for keyword in keywords)
                elif hasattr(x, '__iter__') and not isinstance(x, str):
                    # 对于可迭代对象(不是字符串)，检查每个元素
                    return any(keyword in str(val) for val in x for keyword in keywords)
                else:
                    # 其他类型转为字符串后检查
                    return any(keyword in str(x) for keyword in keywords)
            except Exception:
                # 如果有任何错误，返回False
                return False

        # 查找表头行
        header_rows = []
        for i, row in df.iterrows():
            # 检查行中是否包含任何关键词
            if any(safe_contains_keyword(cell, header_keywords) for cell in row if pd.notna(cell)):
                header_rows.append(i)
        
        if header_in_top and date_in_top:
            header_info["header_position"] = "top"
            header_info["header_rows"] = header_rows
            header_info["is_multi_level"] = len(header_rows) > 1
            header_info["confidence"] = 80 if header_info["is_multi_level"] else 100
        elif header_in_top and date_in_left:
            header_info["header_position"] = "top"
            header_info["header_rows"] = header_rows
            header_info["is_multi_level"] = len(header_rows) > 1
            header_info["confidence"] = 80 if header_info["is_multi_level"] else 100
        elif not header_in_top and date_in_top:
            header_info["header_position"] = "left"
            header_info["header_rows"] = header_rows
            header_info["is_multi_level"] = len(header_rows) > 1
            header_info["confidence"] = 80 if header_info["is_multi_level"] else 100
        else:
            header_info["header_position"] = "left"
            header_info["header_rows"] = header_rows
            header_info["is_multi_level"] = len(header_rows) > 1
            header_info["confidence"] = 80 if header_info["is_multi_level"] else 100
        
        # 更新日志
        for row_idx in header_info["header_rows"]:
            for col_idx in range(df.shape[1]):
                if col_idx < len(df.columns):
                    value = df.iloc[row_idx, col_idx] if row_idx < df.shape[0] else None
                    if pd.notna(value):
                        # 记录字段识别
                        self.excel_logger.log_field_recognition(
                            field_name=str(value),
                            position=(row_idx, col_idx),
                            data_type="header",
                            confidence=header_info["confidence"]
                        )
        
        # 记录表头范围
        if header_info["header_rows"]:
            self.excel_logger.log_data_range(
                df.shape,
                header_range={
                    "start_row": min(header_info["header_rows"]),
                    "end_row": max(header_info["header_rows"]),
                    "start_col": 0,
                    "end_col": df.shape[1] - 1
                }
            )
            
        return header_info
    
    def _extract_dates_from_df(self, df):
        """
        从DataFrame中提取日期信息
        
        Args:
            df: DataFrame数据
            
        Returns:
            list: 提取的日期列表
        """
        dates = []
        date_positions = []
        
        # 日期模式
        date_patterns = [
            r'\d{4}[-/年]\d{1,2}[-/月]',
            r'\d{4}[-/]Q[1-4]',
            r'\d{4}年?第?[一二三四]季度',
            r'\d{4}[-/年]',
            r'\d{4}'
        ]
        
        # 在列名中查找日期
        for col in df.columns:
            if isinstance(col, str) and re.search(r'\d{4}[-/]\d{1,2}', col):
                dates.append(col)
                
        # 在表格内容中查找日期
        if not dates:
            date_pattern = re.compile(r'\d{4}[-/]\d{1,2}')
            
            # 查找顶部行中的日期
            for i in range(min(5, len(df))):
                for col in df.columns:
                    value = df.iloc[i, df.columns.get_loc(col)]
                    if isinstance(value, str) and date_pattern.search(value):
                        dates.append(value)
            
            # 查找第一列中的日期
            first_col = df.iloc[:, 0]
            for value in first_col:
                if isinstance(value, str) and date_pattern.search(value):
                    dates.append(value)
        
        # 提取日期中的年月信息
        clean_dates = []
        for date in dates:
            match = re.search(r'(\d{4})[-/](\d{1,2})', str(date))
            if match:
                year, month = match.groups()
                clean_dates.append(f"{year}-{month.zfill(2)}")
                
        return list(set(clean_dates))
    
    def _parse_multi_level_header(self, df):
        """
        解析多级表头
        
        Args:
            df: 输入的DataFrame
            
        Returns:
            DataFrame: 处理后的DataFrame
        """
        # 检测是否为多级表头
        is_multi_level = False
        header_rows = []
        
        for i in range(min(5, len(df))):
            row = df.iloc[i]
            # 检查行中是否包含科目关键词
            if any(keyword in str(val) for val in row for keyword in 
                   self.balance_sheet_patterns['assets'] + 
                   self.balance_sheet_patterns['liabilities'] + 
                   self.balance_sheet_patterns['equity'] +
                   self.income_statement_patterns['revenue'] +
                   self.income_statement_patterns['costs'] +
                   self.income_statement_patterns['profit']):
                header_rows.append(i)
                is_multi_level = True
        
        if not is_multi_level or not header_rows:
            return df
        
        # 使用多级表头重新解析DataFrame
        max_header_row = max(header_rows)
        columns = []
        
        for i in range(max_header_row + 1):
            row = df.iloc[i]
            row_values = [str(val) if pd.notna(val) else '' for val in row]
            columns.append(row_values)
        
        # 转置并合并列名
        merged_columns = []
        for j in range(len(columns[0])):
            col_parts = []
            for i in range(len(columns)):
                if j < len(columns[i]) and columns[i][j]:
                    col_parts.append(columns[i][j])
            merged_columns.append(' '.join(col_parts))
        
        # 创建新的DataFrame，跳过用作表头的行
        new_df = pd.DataFrame(df.values[max_header_row+1:], columns=merged_columns)
        
        return new_df

    def process_income_statement(self):
        """
        处理损益表
        
        Returns:
            dict: 处理结果
        """
        self._log_processing_info(f"开始处理损益表：{self.income_statement_path}")
        
        try:
            # 使用Excel数据捕获日志记录文件信息
            workbook = load_workbook(self.income_statement_path, data_only=True)
            self.excel_logger.log_file_info(
                self.income_statement_path, 
                sheet_name=workbook.sheetnames[0] if workbook.sheetnames else None,
                sheet_count=len(workbook.sheetnames)
            )
            
            # 尝试不同方式读取Excel
            try:
                df = pd.read_excel(self.income_statement_path)
                self._log_processing_info(f"成功读取Excel文件，数据形状: {df.shape}")
                self.excel_logger.log_data_range(df.shape)
            except Exception as e:
                self._log_processing_error(f"读取Excel文件失败: {e}")
                return {"success": False, "error": f"读取Excel文件失败: {str(e)}"}

            # 检测单位
            unit = self._detect_unit_scale(df)
            self._log_processing_info(f"检测到单位: {unit}")
            
            # 检测表头位置
            layout = self._detect_header_position(df)
            self._log_processing_info(f"检测到的表格布局: 表头位置={layout['header_position']}")
            
            # 处理多级表头
            parsed_df = self._parse_multi_level_header(df)
            
            # 识别合并单元格
            merged_cells = self._identify_merged_cells(self.income_statement_path)
            if merged_cells:
                self._log_processing_info(f"检测到 {len(merged_cells)} 个合并单元格")
                
            # 尝试多种读取方式
            result = None
            successful_patterns = []
            
            # 标准方法
            result, successful_patterns = self._try_process_income_statement(parsed_df, self.income_statement_patterns, unit)
            
            # 如果标准方法失败，尝试跳过前几行
            if not result or not (result['revenue'] or result['costs'] or result['profit']):
                self._log_processing_info("标准方法未能识别到损益表数据，尝试跳过前几行")
                for skip_rows in [1, 2, 3, 5, 7]:
                    self._log_processing_info(f"尝试跳过前 {skip_rows} 行")
                    try:
                        alt_df = pd.read_excel(self.income_statement_path, header=None, skiprows=skip_rows)
                        alt_parsed_df = self._parse_multi_level_header(alt_df)
                        result, successful_patterns = self._try_process_income_statement(alt_parsed_df, self.income_statement_patterns, unit)
                        if result and (result['revenue'] or result['costs'] or result['profit']):
                            self._log_processing_info(f"成功读取损益表，跳过前 {skip_rows} 行")
                            break
                    except Exception as e:
                        self._log_processing_warning(f"尝试跳过前 {skip_rows} 行时出错: {e}")
                        
            # 如果仍然失败，尝试自定义表头
            if not result or not (result['revenue'] or result['costs'] or result['profit']):
                self._log_processing_info("尝试使用自定义表头读取")
                try:
                    custom_df = pd.read_excel(self.income_statement_path, header=None)
                    result, successful_patterns = self._try_process_income_statement_custom(custom_df, unit)
                except Exception as e:
                    self._log_processing_warning(f"尝试使用自定义表头时出错: {e}")
            
            if result and (result['revenue'] or result['costs'] or result['profit']):
                # 验证处理结果
                validation_result = self.data_validator.validate_income_statement(result)
                if not validation_result['is_valid']:
                    self._log_processing_warning("损益表数据验证未通过")
                    for warning in validation_result.get('warnings', []):
                        self._log_processing_warning(warning)
                
                # 保存成功的模板
                if successful_patterns:
                    try:
                        template_id = self.template_store.save_income_statement_template(parsed_df, successful_patterns)
                        self._log_processing_info(f"成功保存损益表模板，ID: {template_id}")
                    except Exception as e:
                        self._log_processing_warning(f"保存模板时出错: {e}")
                
                return result
            else:
                self._log_processing_error("所有方法均未能成功识别损益表数据")
                return {
                    'revenue': {},
                    'costs': {},
                    'profit': {},
                    'dates': []
                }
                
        except Exception as e:
            self._log_processing_error(f"处理损益表时出错: {str(e)}")
            raise
    
    def _process_income_statement_with_template(self, df, template_suggestion):
        """
        使用模板处理损益表
        
        Args:
            df: 输入的DataFrame
            template_suggestion: 模板建议
            
        Returns:
            dict: 处理后的损益表数据
        """
        patterns = template_suggestion.get('patterns', self.income_statement_patterns)
        
        # 检测单位
        unit = self._detect_unit_scale(df)
        
        # 使用模板提供的模式尝试识别
        return self._try_process_income_statement(df, {'start': patterns, 'end': patterns}, unit)[0]
    
    def _try_process_income_statement(self, df, keywords, unit):
        """
        尝试处理损益表
        
        Args:
            df: 输入的DataFrame
            keywords: 关键字字典
            unit: 数值单位
            
        Returns:
            tuple: (处理结果, 成功的匹配模式)
        """
        start_idx, end_idx = self._find_data_range(df, keywords)
        
        if start_idx is None:
            return None, []
            
        # 提取有效数据范围
        data_df = df.iloc[start_idx:end_idx+1]
        
        # 识别日期列
        date_columns = []
        for col_idx, col in enumerate(df.columns):
            col_values = df[col].astype(str)
            date_matches = col_values.str.contains(r'\d{4}[-/]\d{1,2}', regex=True)
            if date_matches.any():
                date_columns.append(col_idx)  # 存储列索引而非列名
        
        # 如果没有找到日期列，尝试识别数值列
        if not date_columns:
            numeric_columns = []
            for col_idx, col in enumerate(df.columns):
                if pd.notna(df[col]).any() and df[col].apply(lambda x: isinstance(x, (int, float)) or 
                                 (isinstance(x, str) and re.match(r'^-?\d+\.?\d*$', str(x).replace(',', '')))).any():
                    numeric_columns.append(col_idx)  # 存储列索引而非列名
            date_columns = numeric_columns[:3]  # 假设前三个数值列为日期列
        
        # 提取收入、成本和利润
        revenue = {}
        costs = {}
        profit = {}
        
        # 提取日期
        dates = []
        for col_idx in date_columns:
            if col_idx < len(df.columns):
                date_value = df.iloc[0, col_idx]
                if pd.notna(date_value):
                    dates.append(str(date_value))
                else:
                    dates.append(f"Column {col_idx}")
        
        # 如果没有从列中找到日期，尝试从表格内容中提取
        if not dates or all(d.startswith("Column") for d in dates):
            extracted_dates = self._extract_dates_from_df(df)
            if extracted_dates:
                dates = extracted_dates
                
        # 确保日期格式一致
        clean_dates = []
        for date in dates:
            match = re.search(r'(\d{4})[-/](\d{1,2})', str(date))
            if match:
                year, month = match.groups()
                clean_dates.append(f"{year}-{month.zfill(2)}")
            else:
                clean_dates.append(date)
        
        dates = clean_dates
        
        # 记录成功匹配的模式
        successful_patterns = []
        
        # 处理数据行
        for idx, row in data_df.iterrows():
            item_name = None
            # 查找第一列作为项目名称
            if len(row) > 0:
                first_col_value = row.iloc[0]  # 使用位置索引
                if pd.notna(first_col_value) and isinstance(first_col_value, str) and len(first_col_value.strip()) > 0:
                    item_name = first_col_value.strip()
            
            if not item_name:
                continue
            
            # 记录成功匹配的项目名称
            if any(keyword in item_name for keyword in keywords.get('revenue', [])):
                successful_patterns.append(item_name)
            elif any(keyword in item_name for keyword in keywords.get('costs', [])):
                successful_patterns.append(item_name)
            elif any(keyword in item_name for keyword in keywords.get('profit', [])):
                successful_patterns.append(item_name)
                
            # 根据关键字确定是收入、成本还是利润
            row_dict = {}
            for i, col_idx in enumerate(date_columns):
                if i < len(dates) and col_idx < len(row):
                    value = row.iloc[col_idx]  # 使用位置索引
                    # 清理和转换数值，同时统一单位
                    cleaned_value = self._clean_numeric_data(value)
                    if not np.isnan(cleaned_value):
                        normalized_value = self._normalize_unit_scale(cleaned_value, unit)
                        row_dict[dates[i]] = normalized_value
            
            if any(keyword in item_name for keyword in ['收入', '营业收入', '主营业务收入']):
                revenue[item_name] = row_dict
            elif any(keyword in item_name for keyword in ['成本', '费用', '营业成本', '主营业务成本']):
                costs[item_name] = row_dict
            elif any(keyword in item_name for keyword in ['利润', '净利润', '毛利', '营业利润']):
                profit[item_name] = row_dict
        
        return {
            'revenue': revenue,
            'costs': costs,
            'profit': profit,
            'dates': dates,
            'unit': unit
        }, successful_patterns
    
    def _try_process_income_statement_custom(self, df, unit):
        """
        使用自定义方法处理复杂的损益表
        
        Args:
            df: 输入的DataFrame
            unit: 数值单位
            
        Returns:
            tuple: (处理结果, 成功的匹配模式)
        """
        revenue = {}
        costs = {}
        profit = {}
        successful_patterns = []
        
        # 提取日期信息
        dates = self._extract_dates_from_df(df)
        
        # 如果没有找到日期，尝试使用列索引
        if not dates:
            numeric_cols = []
            for i, col in enumerate(df.columns):
                if df[col].apply(lambda x: isinstance(x, (int, float)) or 
                              (isinstance(x, str) and re.match(r'^-?\d+\.?\d*$', str(x).replace(',', '')))).any():
                    numeric_cols.append(i)
            
            if numeric_cols:
                dates = [f"Column {col}" for col in numeric_cols[:3]]
        
        # 查找收入、成本和利润的行
        for idx, row in df.iterrows():
            item_found = False
            item_name = None
            
            # 查找第一列是否包含项目名称
            if len(df.columns) > 0:
                first_col_value = row.iloc[0]  # 使用位置索引
                if pd.notna(first_col_value) and isinstance(first_col_value, str):
                    item_name = first_col_value.strip()
                    
                    # 检查是否为收入、成本或利润项目
                    if any(keyword in item_name for keyword in 
                          self.income_statement_patterns['revenue'] + 
                          self.income_statement_patterns['costs'] + 
                          self.income_statement_patterns['profit']):
                        item_found = True
                        successful_patterns.append(item_name)
            
            if not item_found or not item_name:
                continue
                
            # 提取数值
            row_dict = {}
            date_cols = []
            
            # 如果有日期，使用日期作为键
            if dates:
                date_cols = list(range(1, min(len(df.columns), len(dates) + 1)))
                
                for i, col_idx in enumerate(date_cols):
                    if i < len(dates) and col_idx < len(df.columns):
                        value = row.iloc[col_idx]  # 使用位置索引
                        # 清理和转换数值
                        cleaned_value = self._clean_numeric_data(value)
                        if not np.isnan(cleaned_value):
                            normalized_value = self._normalize_unit_scale(cleaned_value, unit)
                            row_dict[dates[i]] = normalized_value
            
            # 根据项目名称分类
            if any(keyword in item_name for keyword in self.income_statement_patterns['revenue']):
                revenue[item_name] = row_dict
            elif any(keyword in item_name for keyword in self.income_statement_patterns['costs']):
                costs[item_name] = row_dict
            elif any(keyword in item_name for keyword in self.income_statement_patterns['profit']):
                profit[item_name] = row_dict
        
        return {
            'revenue': revenue,
            'costs': costs,
            'profit': profit,
            'dates': dates,
            'unit': unit
        }, successful_patterns
    
    def save_processing_logs(self, output_dir=None):
        """
        保存处理日志
        
        Args:
            output_dir: 输出目录，如果为None则使用默认目录
            
        Returns:
            tuple: (资产负债表日志路径, 损益表日志路径)
        """
        # 保存Excel捕获日志
        balance_sheet_log = self.excel_logger.save_log()
        
        # 重置日志以存储损益表信息
        self.excel_logger.reset_log()
        income_statement_log = self.excel_logger.save_log()
        
        return (balance_sheet_log, income_statement_log)
        
    def get_processing_statistics(self):
        """
        获取处理统计信息
        
        Returns:
            dict: 处理统计信息
        """
        # 获取Excel捕获日志的摘要
        capture_summary = self.excel_logger.get_capture_summary()
        
        # 获取处理日志统计
        log_stats = self._get_processing_log_report()
        
        # 合并统计信息
        stats = {
            "file_info": {
                "balance_sheet": os.path.basename(self.balance_sheet_path),
                "income_statement": os.path.basename(self.income_statement_path),
                "company_name": self.company_name
            },
            "capture_stats": capture_summary,
            "processing_logs": log_stats
        }
        
        return stats 