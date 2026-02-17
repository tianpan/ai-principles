import pandas as pd
import re
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def similarity(a, b):
    """Calculate string similarity ratio between two strings."""
    return SequenceMatcher(None, a, b).ratio()

def extract_location(name):
    """
    Extract the location part from a company name.
    Tries to identify Chinese city or region names at the beginning of company names.
    """
    # Common province/city/region name patterns in company names
    location_patterns = [
        r'^([\u4e00-\u9fa5]{2,4}?)(市|省|自治区|特别行政区|区|县|镇)?',  # General location pattern
        r'^([\u4e00-\u9fa5]{2,4})',  # Just get first 2-4 characters as potential location
    ]
    
    for pattern in location_patterns:
        match = re.search(pattern, name)
        if match:
            return match.group(1)
    
    return ""

def clean_name(name):
    """Clean up company names for better matching."""
    name = name.replace('有限公司', '').replace('有限责任公司', '')
    name = name.replace('股份', '').replace('责任', '')
    name = name.replace('(', '').replace(')', '')
    name = name.replace('（', '').replace('）', '')
    name = re.sub(r'\s+', '', name)
    return name

def detect_encoding_and_read(filename):
    """Try to detect encoding and read the file content."""
    encodings = ['utf-8', 'gbk', 'gb18030', 'latin1']
    
    # Read file in binary mode
    with open(filename, 'rb') as f:
        content = f.read()
    
    # Try different encodings
    for encoding in encodings:
        try:
            decoded = content.decode(encoding)
            lines = decoded.split('\n')
            print(f"Successfully decoded {filename} with {encoding}")
            return lines, encoding
        except UnicodeDecodeError:
            print(f"Failed to decode {filename} with {encoding}")
    
    print(f"Could not decode {filename} with any of the attempted encodings")
    return None, None

def parse_csv_lines(lines, file_num):
    """Parse CSV lines and extract company names."""
    companies = []
    
    for line in lines[1:]:  # Skip header
        if not line.strip():
            continue
        
        parts = line.split(',')
        if file_num == 1 and len(parts) >= 2:
            company = parts[1].strip()
            if company:
                companies.append(company)
        elif file_num == 2 and len(parts) >= 3:
            company = parts[2].strip()
            if company:
                companies.append(company)
    
    return companies

def advanced_similarity_comparison(name1, name2):
    """
    More advanced similarity comparison that considers location differences.
    Returns a tuple (is_match, similarity_score)
    """
    # Initial clean for basic comparison
    clean1 = clean_name(name1)
    clean2 = clean_name(name2)
    
    # Get basic similarity
    sim = similarity(clean1, clean2)
    
    # Extract potential locations
    loc1 = extract_location(name1)
    loc2 = extract_location(name2)
    
    # If locations are different and both are present, reduce similarity score
    if loc1 and loc2 and loc1 != loc2:
        # Different locations should be a strong signal that these are different entities
        return False, sim * 0.6  # Penalize the similarity score
    
    # Remove location parts and see if the remaining parts are very similar
    rest1 = clean1.replace(loc1, '', 1) if loc1 else clean1
    rest2 = clean2.replace(loc2, '', 1) if loc2 else clean2
    
    # If the remaining parts (after removing location) are very similar,
    # and the locations are different, it's likely different branches of same company
    remaining_sim = similarity(rest1, rest2)
    
    # If extremely similar naming pattern but different locations, it's usually different companies
    if remaining_sim > 0.9 and loc1 and loc2 and loc1 != loc2:
        # These are likely different entities with similar naming patterns (e.g., different branches)
        is_match = False
    else:
        # Use a threshold for determining matches
        is_match = sim > 0.75
    
    return is_match, sim

# Read and process the files
print("Reading file 1...")
lines1, encoding1 = detect_encoding_and_read('1.板块收集.csv')
if not lines1:
    print("Failed to read file 1")
    exit(1)

print("Reading file 2...")
lines2, encoding2 = detect_encoding_and_read('2.TA800.csv')
if not lines2:
    print("Failed to read file 2")
    exit(1)

# Extract company names
companies1 = parse_csv_lines(lines1, 1)
companies2 = parse_csv_lines(lines2, 2)

print(f"Extracted {len(companies1)} companies from file 1")
print(f"Extracted {len(companies2)} companies from file 2")

# Find exact matches
exact_matches = []
for i, name1 in enumerate(companies1):
    if name1 in companies2:
        exact_matches.append((name1, name1))

print(f"Found {len(exact_matches)} exact matches")

# Find similar matches (potential matches) using the advanced comparison
threshold = 0.75  # Base threshold, the advanced comparison may override this
similar_matches = []
for i, name1 in enumerate(companies1):
    for j, name2 in enumerate(companies2):
        if name1 != name2:  # Skip exact matches
            is_match, sim = advanced_similarity_comparison(name1, name2)
            if is_match and sim > threshold:
                similar_matches.append((name1, name2, sim))

# Sort by similarity
similar_matches.sort(key=lambda x: x[2], reverse=True)
print(f"Found {len(similar_matches)} potential matches")

# Find unmatched companies
not_matched_file1 = []
matched_names1 = set([m[0] for m in exact_matches] + [m[0] for m in similar_matches])
for name in companies1:
    if name not in matched_names1:
        not_matched_file1.append(name)

not_matched_file2 = []
matched_names2 = set([m[1] for m in exact_matches] + [m[1] for m in similar_matches])
for name in companies2:
    if name not in matched_names2:
        not_matched_file2.append(name)

print(f"Found {len(not_matched_file1)} unmatched companies in file 1")
print(f"Found {len(not_matched_file2)} unmatched companies in file 2")

# Create Excel file
wb = Workbook()

# Define styles
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Orange
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # Red
header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
bold_font = Font(bold=True)
border = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'), 
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)

# Create worksheets
ws_exact = wb.active
ws_exact.title = "精确匹配"
ws_similar = wb.create_sheet(title="可能匹配")
ws_unmatched1 = wb.create_sheet(title="文件1未匹配")
ws_unmatched2 = wb.create_sheet(title="文件2未匹配")
ws_summary = wb.create_sheet(title="汇总信息")

# Set up exact matches sheet
ws_exact.append(["序号", "文件1中的公司", "文件2中的公司", "状态"])
for row in ws_exact.iter_rows(min_row=1, max_row=1, min_col=1, max_col=4):
    for cell in row:
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = border

for i, (name1, name2) in enumerate(exact_matches, 1):
    row = [i, name1, name2, "精确匹配"]
    ws_exact.append(row)
    for cell in ws_exact[ws_exact.max_row]:
        cell.fill = green_fill
        cell.border = border

# Set up similar matches sheet
ws_similar.append(["序号", "文件1中的公司", "文件2中的公司", "相似度", "状态", "地名差异"])
for row in ws_similar.iter_rows(min_row=1, max_row=1, min_col=1, max_col=6):
    for cell in row:
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = border

for i, (name1, name2, sim) in enumerate(similar_matches, 1):
    loc1 = extract_location(name1)
    loc2 = extract_location(name2)
    location_diff = "不同地区" if loc1 and loc2 and loc1 != loc2 else ""
    
    row = [i, name1, name2, f"{sim:.2f}", "可能匹配", location_diff]
    ws_similar.append(row)
    for cell in ws_similar[ws_similar.max_row]:
        cell.fill = orange_fill
        cell.border = border

# Set up unmatched companies from file 1
ws_unmatched1.append(["序号", "文件1中的公司", "状态"])
for row in ws_unmatched1.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3):
    for cell in row:
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = border

for i, name in enumerate(not_matched_file1, 1):
    row = [i, name, "未匹配"]
    ws_unmatched1.append(row)
    for cell in ws_unmatched1[ws_unmatched1.max_row]:
        cell.fill = red_fill
        cell.border = border

# Set up unmatched companies from file 2
ws_unmatched2.append(["序号", "文件2中的公司", "状态"])
for row in ws_unmatched2.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3):
    for cell in row:
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = border

for i, name in enumerate(not_matched_file2, 1):
    row = [i, name, "未匹配"]
    ws_unmatched2.append(row)
    for cell in ws_unmatched2[ws_unmatched2.max_row]:
        cell.fill = red_fill
        cell.border = border

# Add data statistics summary sheet
ws_summary.append(["统计信息"])
ws_summary.append(["文件1总公司数", len(companies1)])
ws_summary.append(["文件2总公司数", len(companies2)])
ws_summary.append(["精确匹配公司数", len(exact_matches)])
ws_summary.append(["可能匹配公司数", len(similar_matches)])
ws_summary.append(["文件1未匹配公司数", len(not_matched_file1)])
ws_summary.append(["文件2未匹配公司数", len(not_matched_file2)])

for row in ws_summary.iter_rows(min_row=1, max_row=1, min_col=1, max_col=1):
    for cell in row:
        cell.font = bold_font

# Adjust column widths
for ws in [ws_exact, ws_similar, ws_unmatched1, ws_unmatched2, ws_summary]:
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

# Save the workbook
excel_file = "企业匹配结果_优化版.xlsx"
wb.save(excel_file)
print(f"Excel文件已创建: {excel_file}") 